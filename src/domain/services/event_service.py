"""Domain service: EventService — event logging, retrieval, and Excel export."""

import io
from datetime import datetime, timezone
from uuid import UUID, uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..models.event import Event, EventType
from ..ports.event_repository import EventFilters, EventPage, IEventRepository


class EventService:
    """Manages event logging, retrieval, and Excel export.

    Wraps the IEventRepository port with higher-level helpers so that
    callers don't need to construct raw Event objects or worry about the
    details of the export format.

    Args:
        event_repo: Repository for persisting and querying audit events.
    """

    def __init__(self, event_repo: IEventRepository) -> None:
        self._repo = event_repo

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def log(
        self,
        event_type: EventType,
        description: str,
        metadata: dict | None = None,
        user_id: UUID | None = None,
    ) -> Event:
        """Create and persist an audit-log event.

        Args:
            event_type: Category of the action being recorded.
            description: Human-readable summary of what happened.
            metadata: Optional structured key-value payload for additional context.
            user_id: UUID of the user who triggered the event; ``None`` for system events.

        Returns:
            The persisted :class:`Event` entity (with DB-assigned fields populated).
        """
        event = Event(
            id=uuid4(),
            event_type=event_type,
            description=description,
            metadata=metadata or {},
            user_id=user_id,
            created_at=datetime.now(timezone.utc),
        )
        return self._repo.log_event(event)

    def list_events(self, filters: EventFilters) -> EventPage:
        """Return a paginated, filtered list of audit events.

        Args:
            filters: Query parameters including event type, date range,
                description substring, and pagination settings.

        Returns:
            An :class:`EventPage` containing the matching items and
            pagination metadata (total count, current page, page size).
        """
        return self._repo.get_events(filters)

    def export_to_excel(self, filters: EventFilters) -> bytes:
        """Export filtered events to an Excel (.xlsx) file.

        Retrieves all matching events (unpaginated), writes them to a
        formatted workbook with a styled header row, alternating row
        colours, frozen pane, and auto-width columns.

        Args:
            filters: Filter criteria for selecting events to export
                (pagination fields are ignored — all matching events are included).

        Returns:
            Raw bytes of the generated ``.xlsx`` file, ready for streaming
            in an HTTP response.
        """
        events = self._repo.get_events_for_export(filters)

        wb = Workbook()
        ws = wb.active
        ws.title = "Eventos"

        # Header row
        headers = ["ID", "Tipo", "Descripción", "Usuario ID", "Fecha y Hora"]
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        ws.freeze_panes = "A2"

        # Data rows with alternating background colours
        even_fill = PatternFill(
            start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
        )
        for row_idx, event in enumerate(events, 2):
            row_data = [
                str(event.id),
                event.event_type.value,
                event.description,
                str(event.user_id) if event.user_id is not None else "",
                event.created_at.isoformat(),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        # Auto-fit column widths (capped at 50 characters)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
