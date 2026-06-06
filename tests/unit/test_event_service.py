"""Unit tests for EventService — log, list_events, and export_to_excel."""

import io
import pytest
from datetime import datetime, timezone
from uuid import uuid4

import openpyxl

from src.domain.models.event import Event, EventType
from src.domain.ports.event_repository import EventFilters, EventPage
from src.domain.services.event_service import EventService

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_service(mock_event_repo) -> EventService:
    """Instantiate EventService with an injected mock repository."""
    return EventService(mock_event_repo)


def _make_event(event_type: EventType, description: str = "test", user_id=None) -> Event:
    """Build a minimal Event entity for use in test stubs."""
    return Event(
        id=uuid4(),
        event_type=event_type,
        description=description,
        metadata={},
        user_id=user_id,
        created_at=datetime.now(timezone.utc),
    )


# ---------------------------------------------------------------------------
# Tests: log
# ---------------------------------------------------------------------------


class TestEventServiceLog:
    """Tests for EventService.log."""

    @pytest.mark.parametrize("event_type", list(EventType))
    def test_log_all_event_types(self, event_type, mock_event_repo):
        """Every EventType value must be accepted and persisted via the repository."""
        service = _make_service(mock_event_repo)

        event = service.log(event_type, "test description")

        assert event.event_type == event_type
        mock_event_repo.log_event.assert_called_once()

    def test_log_returns_event_with_id(self, mock_event_repo):
        """The returned Event must have a non-None UUID id field."""
        service = _make_service(mock_event_repo)

        event = service.log(EventType.CSV_UPLOAD, "Upload complete")

        assert event.id is not None

    def test_log_calls_repo_log_event(self, mock_event_repo):
        """log() must delegate persistence to IEventRepository.log_event exactly once."""
        service = _make_service(mock_event_repo)

        service.log(EventType.USER_LOGIN, "User logged in")

        mock_event_repo.log_event.assert_called_once()

    def test_log_with_user_id(self, mock_event_repo):
        """When user_id is provided, the persisted event must carry it."""
        service = _make_service(mock_event_repo)
        user = uuid4()

        event = service.log(EventType.CSV_UPLOAD, "Upload", user_id=user)

        assert event.user_id == user

    def test_log_without_user_id(self, mock_event_repo):
        """When user_id is omitted, the persisted event must have user_id=None."""
        service = _make_service(mock_event_repo)

        event = service.log(EventType.ERROR, "System error")

        assert event.user_id is None

    def test_log_stores_metadata(self, mock_event_repo):
        """Provided metadata must be preserved on the returned event."""
        service = _make_service(mock_event_repo)
        meta = {"rows": 100, "bucket": "my-bucket"}

        event = service.log(EventType.CSV_UPLOAD, "Upload", metadata=meta)

        assert event.metadata == meta

    def test_log_empty_metadata_defaults_to_empty_dict(self, mock_event_repo):
        """When metadata is not supplied, it must default to an empty dict (not None)."""
        service = _make_service(mock_event_repo)

        event = service.log(EventType.TOKEN_RENEWAL, "Renewed")

        assert event.metadata == {}

    def test_log_preserves_description_text(self, mock_event_repo):
        """The event description must be preserved exactly as provided."""
        service = _make_service(mock_event_repo)
        description = "CSV processed with 5 warnings"

        event = service.log(EventType.CSV_VALIDATION, description)

        assert event.description == description

    def test_log_sets_timezone_aware_created_at(self, mock_event_repo):
        """Created timestamp must be timezone-aware."""
        service = _make_service(mock_event_repo)

        event = service.log(EventType.USER_LOGIN, "Login")

        assert event.created_at.tzinfo is not None

    def test_log_returns_repository_result(self, mock_event_repo):
        """Service should return exactly what repository returns from log_event."""
        service = _make_service(mock_event_repo)
        persisted = _make_event(EventType.ERROR, "Persisted")
        mock_event_repo.log_event.return_value = persisted

        result = service.log(EventType.ERROR, "Original")

        assert result is persisted


# ---------------------------------------------------------------------------
# Tests: list_events
# ---------------------------------------------------------------------------


class TestEventServiceListEvents:
    """Tests for EventService.list_events."""

    def test_list_returns_event_page(self, mock_event_repo):
        """list_events must return the EventPage produced by the repository."""
        expected_page = EventPage(items=[], total=0, page=1, page_size=50)
        mock_event_repo.get_events.return_value = expected_page
        service = _make_service(mock_event_repo)

        result = service.list_events(EventFilters())

        assert result is expected_page

    def test_list_passes_filters_to_repo(self, mock_event_repo):
        """The exact EventFilters object must be forwarded to get_events."""
        mock_event_repo.get_events.return_value = EventPage(
            items=[], total=0, page=1, page_size=10
        )
        service = _make_service(mock_event_repo)
        filters = EventFilters(event_type=EventType.CSV_UPLOAD, page=2, page_size=10)

        service.list_events(filters)

        mock_event_repo.get_events.assert_called_once_with(filters)

    def test_list_returns_items_from_repo(self, mock_event_repo):
        """Items in the returned EventPage must match what the repository provided."""
        events = [_make_event(EventType.CSV_UPLOAD, f"upload {i}") for i in range(3)]
        mock_event_repo.get_events.return_value = EventPage(
            items=events, total=3, page=1, page_size=50
        )
        service = _make_service(mock_event_repo)

        result = service.list_events(EventFilters())

        assert len(result.items) == 3

    def test_list_preserves_pagination_fields(self, mock_event_repo):
        """Page metadata must be returned unchanged from repository output."""
        expected = EventPage(items=[], total=15, page=2, page_size=5)
        mock_event_repo.get_events.return_value = expected
        service = _make_service(mock_event_repo)

        result = service.list_events(EventFilters(page=2, page_size=5))

        assert result.total == 15
        assert result.page == 2
        assert result.page_size == 5

    def test_list_forwards_event_type_filter(self, mock_event_repo):
        """Event type filter must be forwarded unchanged."""
        mock_event_repo.get_events.return_value = EventPage(items=[], total=0, page=1, page_size=50)
        service = _make_service(mock_event_repo)
        filters = EventFilters(event_type=EventType.AI_EXTRACTION)

        service.list_events(filters)

        sent_filters = mock_event_repo.get_events.call_args.args[0]
        assert sent_filters.event_type == EventType.AI_EXTRACTION

    def test_list_forwards_description_filter(self, mock_event_repo):
        """Description substring filter must be forwarded unchanged."""
        mock_event_repo.get_events.return_value = EventPage(items=[], total=0, page=1, page_size=50)
        service = _make_service(mock_event_repo)
        filters = EventFilters(description_contains="csv")

        service.list_events(filters)

        sent_filters = mock_event_repo.get_events.call_args.args[0]
        assert sent_filters.description_contains == "csv"

    def test_list_forwards_date_range_filters(self, mock_event_repo):
        """Date range filters must be forwarded unchanged."""
        now = datetime.now(timezone.utc)
        filters = EventFilters(date_from=now, date_to=now)
        mock_event_repo.get_events.return_value = EventPage(items=[], total=0, page=1, page_size=50)
        service = _make_service(mock_event_repo)

        service.list_events(filters)

        sent_filters = mock_event_repo.get_events.call_args.args[0]
        assert sent_filters.date_from == now
        assert sent_filters.date_to == now

    def test_list_empty_result_returns_no_items(self, mock_event_repo):
        """Empty repository page must return an empty items list."""
        mock_event_repo.get_events.return_value = EventPage(items=[], total=0, page=1, page_size=50)
        service = _make_service(mock_event_repo)

        result = service.list_events(EventFilters())

        assert result.items == []

    def test_list_propagates_repository_errors(self, mock_event_repo):
        """Repository exceptions should bubble up to caller."""
        mock_event_repo.get_events.side_effect = RuntimeError("db unavailable")
        service = _make_service(mock_event_repo)

        with pytest.raises(RuntimeError, match="db unavailable"):
            service.list_events(EventFilters())


# ---------------------------------------------------------------------------
# Tests: export_to_excel
# ---------------------------------------------------------------------------


class TestEventServiceExportToExcel:
    """Tests for EventService.export_to_excel."""

    def test_export_returns_bytes(self, mock_event_repo):
        """export_to_excel must return a non-empty bytes object."""
        mock_event_repo.get_events_for_export.return_value = []
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_creates_valid_xlsx(self, mock_event_repo):
        """The returned bytes must be a loadable .xlsx workbook with the expected headers."""
        mock_event_repo.get_events_for_export.return_value = []
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, 6)]
        assert headers == ["ID", "Tipo", "Descripción", "Usuario ID", "Fecha y Hora"]

    def test_export_includes_all_events(self, mock_event_repo):
        """Workbook must contain one data row per event plus one header row."""
        events = [_make_event(EventType.CSV_UPLOAD, f"event {i}") for i in range(5)]
        mock_event_repo.get_events_for_export.return_value = events
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.max_row == 6  # 1 header + 5 data rows

    def test_export_with_no_events_returns_empty_xlsx(self, mock_event_repo):
        """When there are no events, the workbook must contain only the header row."""
        mock_event_repo.get_events_for_export.return_value = []
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.max_row == 1

    def test_export_passes_filters_to_repo(self, mock_event_repo):
        """The EventFilters must be forwarded to get_events_for_export."""
        mock_event_repo.get_events_for_export.return_value = []
        service = _make_service(mock_event_repo)
        filters = EventFilters(event_type=EventType.ERROR)

        service.export_to_excel(filters)

        mock_event_repo.get_events_for_export.assert_called_once_with(filters)

    def test_export_row_data_matches_event_fields(self, mock_event_repo):
        """Each data row in the workbook must reflect the corresponding event's fields."""
        user = uuid4()
        events = [_make_event(EventType.USER_LOGIN, "Login event", user_id=user)]
        mock_event_repo.get_events_for_export.return_value = events
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # Row 2 is the first data row
        assert ws.cell(2, 2).value == EventType.USER_LOGIN.value
        assert ws.cell(2, 3).value == "Login event"
        assert ws.cell(2, 4).value == str(user)

    def test_export_anonymous_event_has_empty_user_id_cell(self, mock_event_repo):
        """Events without a user_id must produce an empty string in the User ID column."""
        events = [_make_event(EventType.ERROR, "System error", user_id=None)]
        mock_event_repo.get_events_for_export.return_value = events
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # openpyxl stores empty-string cells as None; both are equivalent here
        assert ws.cell(2, 4).value in ("", None)

    def test_export_sheet_name_is_eventos(self, mock_event_repo):
        """Workbook active sheet should be named 'Eventos'."""
        mock_event_repo.get_events_for_export.return_value = []
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb.active.title == "Eventos"

    def test_export_freezes_header_row(self, mock_event_repo):
        """Workbook must freeze pane below the header row."""
        mock_event_repo.get_events_for_export.return_value = []
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb.active.freeze_panes == "A2"

    def test_export_columns_have_reasonable_width_cap(self, mock_event_repo):
        """Auto-sized columns should be capped at <= 50 width."""
        long_desc = "x" * 400
        events = [_make_event(EventType.ERROR, long_desc)]
        mock_event_repo.get_events_for_export.return_value = events
        service = _make_service(mock_event_repo)

        result = service.export_to_excel(EventFilters())

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        for col in ["A", "B", "C", "D", "E"]:
            width = ws.column_dimensions[col].width
            assert width is not None
            assert width <= 50
